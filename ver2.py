import sys
from inspect import signature

from PyQt5.QtCore import QDateTime
from PyQt5.QtWidgets import QApplication, QWidget, QVBoxLayout, QHBoxLayout, QLabel, QLineEdit, QPushButton, \
    QTableWidget, QTableWidgetItem, QGridLayout, QHeaderView, QMessageBox, QDialog, QButtonGroup, \
    QRadioButton, QComboBox, QFileDialog, QDialogButtonBox, QDateTimeEdit
from openpyxl.styles import Font, Alignment


def pt_to_col_width(pt):
    return pt / 6.2857  # Преобразование пунктов в ширину столбца


class SignaturesDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle('Подписи')
        self.setGeometry(200, 200, 500, 200)
        layout = QVBoxLayout()
        form_layout = QVBoxLayout()

        input_width = 350  # Ширина полей ввода в форме подписи

        self.workers = self.load_workers('workers.txt')


        head_layout = QHBoxLayout()
        head_layout.addWidget(QLabel('Заведующий:'))
        self.head_fio = QComboBox()
        self.head_fio.setEditable(True)
        self.head_fio.setFixedWidth(input_width)
        head_layout.addWidget(self.head_fio)
        head_layout.addWidget(QLabel('Должность:'))
        self.head_position = QComboBox()
        self.head_position.addItems(['Главный заведующий', 'Помощник заведующего'])
        self.head_position.setFixedWidth(input_width)
        self.head_position.currentTextChanged.connect(self.update_head_fio)
        head_layout.addWidget(self.head_position)
        form_layout.addLayout(head_layout)

        accountant_layout = QHBoxLayout()
        accountant_layout.addWidget(QLabel('Бухгалтер:'))
        self.accountant_fio = QComboBox()
        self.accountant_fio.setEditable(True)
        self.accountant_fio.setFixedWidth(input_width)
        accountant_layout.addWidget(self.accountant_fio)
        accountant_layout.addWidget(QLabel('Должность:'))
        self.accountant_position = QComboBox()
        self.accountant_position.addItems(['Бухгалтер', 'Помощник бухгалтера'])
        self.accountant_position.setFixedWidth(input_width)
        self.accountant_position.currentTextChanged.connect(self.update_accountant_fio)
        accountant_layout.addWidget(self.accountant_position)
        form_layout.addLayout(accountant_layout)

        approve_layout = QHBoxLayout()
        approve_layout.addWidget(QLabel('Утверждаю:'))
        self.approve_fio = QComboBox()
        self.approve_fio.setEditable(True)
        self.approve_fio.setFixedWidth(input_width)
        approve_layout.addWidget(self.approve_fio)
        approve_layout.addWidget(QLabel('Должность:'))
        self.approve_position = QComboBox()
        self.approve_position.addItems(['Главный директор', 'Заместитель директора'])
        self.approve_position.setFixedWidth(input_width)
        self.approve_position.currentTextChanged.connect(self.update_approve_fio)
        approve_layout.addWidget(self.approve_position)
        form_layout.addLayout(approve_layout)

        layout.addLayout(form_layout)

        button_box = QDialogButtonBox(QDialogButtonBox.Save | QDialogButtonBox.Cancel)
        button_box.button(QDialogButtonBox.Save).setText('Сохранить')
        button_box.button(QDialogButtonBox.Cancel).setText('Отмена')
        button_box.accepted.connect(self.accept)
        button_box.rejected.connect(self.reject)
        layout.addWidget(button_box)

        self.setLayout(layout)

    def load_workers(self, filename):
        """Загружает данные из файла workers.txt."""
        workers = {}
        with open(filename, 'r', encoding='utf-8') as file:
            for line in file:
                position, fio = line.strip().split(': ')
                if position not in workers:
                    workers[position] = []
                workers[position].append(fio)
        return workers

    def update_head_fio(self):
        """Обновляет список ФИО для заведующего."""
        self.head_fio.clear()
        position = self.head_position.currentText()
        if position in self.workers:
            self.head_fio.addItems(self.workers[position])

    def update_accountant_fio(self):
        """Обновляет список ФИО для бухгалтера."""
        self.accountant_fio.clear()
        position = self.accountant_position.currentText()
        if position in self.workers:
            self.accountant_fio.addItems(self.workers[position])

    def update_approve_fio(self):
        """Обновляет список ФИО для утверждающего."""
        self.approve_fio.clear()
        position = self.approve_position.currentText()
        if position in self.workers:
            self.approve_fio.addItems(self.workers[position])

    def get_signatures(self):
        """Возвращает введенные подписи."""
        return {
            'head': {'fio': self.head_fio.currentText(), 'position': self.head_position.currentText()},
            'accountant': {'fio': self.accountant_fio.currentText(),
                           'position': self.accountant_position.currentText()},
            'approve': {'fio': self.approve_fio.currentText(), 'position': self.approve_position.currentText()}
        }



class OP1Form(QWidget):
    def __init__(self):
        super().__init__()
        self.org_combobox = None
        self.dish_combobox = None
        self.signatures_data = None  # Атрибут для хранения данных подписей - чет не работает
        self.date_time_edit = None
        self.name_org = None
        self.operation_combobox = None
        self.initUI()

    def initUI(self):
        self.setWindowTitle('Унифицированная форма ОП-1')
        self.setGeometry(100, 100, 1400, 1200)

        main_layout = QVBoxLayout()

        # Заголовок
        title = QLabel('Унифицированная форма ОП-1')
        title.setStyleSheet('font-size: 24px; font-weight: bold;')
        main_layout.addWidget(title)

        grid_layout = QGridLayout()

        self.operation_combobox = self.add_searchable_combobox(grid_layout, 'Вид операции:', ['Приготовление', 'Обработка', 'Упаковка'], 0, 0)

        self.name_org = self.add_line_edit(grid_layout, 'Номер:', '0000-000519', 1, 0)
        self.date_time_edit = self.add_date_time(grid_layout, 'От:', '23.10.2020', 1, 2)

        self.add_line_edit(grid_layout, 'Составлено на:', '1,000', 0, 6)
        self.add_line_edit(grid_layout, 'Выход для печати:', '100', 1, 6)
        self.price_input = QLineEdit('0,00')
        grid_layout.addWidget(QLabel('Цена блюда:'), 2, 6)
        grid_layout.addWidget(self.price_input, 2, 7)

        self.org_combobox = self.add_searchable_combobox(grid_layout, 'Организация:',
                                                         ['Организация 1', 'Организация 2', 'Организация 3'], 0, 2)
        self.dish_combobox = self.add_searchable_combobox(grid_layout, 'Номенклатура:',
                                                          ['Салат мясной', 'Салат овощной', 'Гречка с тушенкой'], 2, 0)

        main_layout.addLayout(grid_layout)

        table_label = QLabel('Ингредиенты:')
        table_label.setStyleSheet('font-size: 12px; font-weight: bold;')

        table_header_layout = QHBoxLayout()
        table_header_layout.addWidget(table_label)
        table_header_layout.addStretch()

        add_button = QPushButton('Добавить')
        add_button.setStyleSheet('font-size: 12px;')
        add_button.clicked.connect(self.add_row_to_table)
        table_header_layout.addWidget(add_button)

        main_layout.addLayout(table_header_layout)

        self.table = QTableWidget()
        self.table.setColumnCount(10)
        self.table.setHorizontalHeaderLabels(
            ['№ п/п', 'Наименование', 'Код', 'Единица', 'Код ОКЕИ', 'Цена, руб. коп.',
             'Норма брутто', '% Потерь', 'Норма нетто', 'Сумма, руб. коп.'])

        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.table.horizontalHeader().setStretchLastSection(True)

        row_height = 20
        for i in range(self.table.rowCount()):
            self.table.setRowHeight(i, row_height)

        main_layout.addWidget(self.table)

        responsible_comment_layout = QHBoxLayout()

        # Ответственный
        responsible_label = QLabel('Ответственный:')
        responsible_label.setStyleSheet('font-size: 12px;')
        self.responsible_input = QComboBox()
        self.responsible_input.addItem('Администратор')
        self.responsible_input.addItem('Менеджер')

        responsible_comment_layout.addWidget(responsible_label)
        responsible_comment_layout.addWidget(self.responsible_input)

        comment_label = QLabel('Комментарий:')
        comment_label.setStyleSheet('font-size: 12px;')
        self.comment_input = QLineEdit()
        self.comment_input.setFixedHeight(20)

        responsible_comment_layout.addWidget(comment_label)
        responsible_comment_layout.addWidget(self.comment_input)

        main_layout.addLayout(responsible_comment_layout)

        button_layout = QHBoxLayout()
        save_button = QPushButton('Загрузить')
        save_button.setStyleSheet('font-size: 12px;')
        save_button.clicked.connect(self.show_save_dialog)
        print_button = QPushButton('Печать')
        print_button.setStyleSheet('font-size: 12px;')
        print_button.clicked.connect(self.show_print_table)
        button_layout.addWidget(save_button)
        button_layout.addWidget(print_button)
        main_layout.addLayout(button_layout)

        signatures_button = QPushButton('Подписи')
        signatures_button.setStyleSheet('font-size: 12px;')
        signatures_button.clicked.connect(self.show_signatures_dialog)
        signatures_button.setFixedSize(100, 30)
        main_layout.addWidget(signatures_button)

        self.setLayout(main_layout)

        # Загрузка данных из файлов
        self.load_data()

        # Подключение сигнала cellChanged
        self.table.cellChanged.connect(self.on_cell_changed)

    def load_data(self):
        """Загружает данные из файлов data.txt и code.txt."""
        self.name_to_code = {}
        with open('data.txt', 'r', encoding='utf-8') as file:
            for line in file:
                code, name = line.strip().split(':')
                self.name_to_code[name] = code

        self.unit_to_okei = {}
        with open('code.txt', 'r', encoding='utf-8') as file:
            for line in file:
                unit, okei = line.strip().split(': ')
                self.unit_to_okei[unit] = okei



    def load_operations_codes(self, filename='operations_code.txt'):
        """Загружает соответствие операций и кодов из файла."""
        operations_codes = {}
        with open(filename, 'r', encoding='utf-8') as file:
            for line in file:
                operation, code = line.strip().split(':')
                operations_codes[operation] = code
        return operations_codes

    def add_date_time(self, layout, label_text, default_date_time, row, column):
        date_time_edit = QDateTimeEdit()
        date_time_edit.setDisplayFormat("dd.MM.yyyy")
        date_time_edit.setDateTime(QDateTime.fromString(default_date_time, "dd.MM.yyyy"))

        layout.addWidget(QLabel(label_text), row, column)
        layout.addWidget(date_time_edit, row, column + 1)

        return date_time_edit

    def get_date(self):
        selected_date = self.date_time_edit.dateTime().toString("dd.MM.yyyy")
        return selected_date

    def get_number_value(self):
        return self.name_org.text()

    def show_signatures_dialog(self):
        if self.signatures_data is None:  # Если данные еще не сохранены
            dialog = SignaturesDialog(self)
            if dialog.exec_() == QDialog.Accepted:
                self.signatures_data = dialog.get_signatures()  # Сохраняем данные
                QMessageBox.information(self, "Подписи",
                                        f"Заведующий: {self.signatures_data['head']['fio']} ({self.signatures_data['head']['position']})\n"
                                        f"Бухгалтер: {self.signatures_data['accountant']['fio']} ({self.signatures_data['accountant']['position']})\n"
                                        f"Утверждаю: {self.signatures_data['approve']['fio']} ({self.signatures_data['approve']['position']})")
            else:
                self.signatures_data = {}
        else:
            QMessageBox.information(self, "Подписи",
                                    f"Заведующий: {self.signatures_data['head']['fio']} ({self.signatures_data['head']['position']})\n"
                                    f"Бухгалтер: {self.signatures_data['accountant']['fio']} ({self.signatures_data['accountant']['position']})\n"
                                    f"Утверждаю: {self.signatures_data['approve']['fio']} ({self.signatures_data['approve']['position']})")

    def add_line_edit(self, layout, label_text, default_text='', row=0, col=0):
        label = QLabel(label_text)
        label.setStyleSheet('font-size: 12px;')
        line_edit = QLineEdit(default_text)
        layout.addWidget(label, row, col)
        layout.addWidget(line_edit, row, col + 1)
        return line_edit

    def add_searchable_combobox(self, layout, label_text, items, row=0, col=0):
        label = QLabel(label_text)
        label.setStyleSheet('font-size: 12px;')
        combo_box = QComboBox()
        combo_box.setEditable(True)
        combo_box.addItems(items)
        combo_box.setCurrentIndex(-1)
        layout.addWidget(label, row, col)
        layout.addWidget(combo_box, row, col + 1)
        return combo_box

    def show_save_dialog(self):
        self.save_dialog = SaveDialog(self)
        self.save_dialog.show()

    def show_print_table(self):
        data = []
        for row in range(self.table.rowCount()):
            row_data = []
            for col in range(self.table.columnCount()):
                item = self.table.item(row, col)
                row_data.append(item.text() if item else "")
            data.append(row_data)

        print_dialog = PrintDialog(data, self)
        print_dialog.exec_()

    def add_row_to_table(self):
        """Добавляет новую строку в таблицу."""
        row_count = self.table.rowCount()
        self.table.insertRow(row_count)
        self.table.setItem(row_count, 0, QTableWidgetItem(str(row_count + 1)))  # Автоматическое заполнение № п/п

        # Наименование
        name_item = QTableWidgetItem()
        self.table.setItem(row_count, 1, name_item)

        # Код
        code_item = QTableWidgetItem()
        self.table.setItem(row_count, 2, code_item)

        # Единица измерения
        unit_combo = QComboBox()
        unit_combo.addItems(["кг", "г", "шт"])
        unit_combo.currentTextChanged.connect(lambda text, row=row_count: self.update_okei_code(row))
        self.table.setCellWidget(row_count, 3, unit_combo)

        # Код ОКЕИ
        okei_item = QTableWidgetItem()
        self.table.setItem(row_count, 4, okei_item)

        # Остальные поля
        for col in range(5, self.table.columnCount()):
            self.table.setItem(row_count, col, QTableWidgetItem(""))

    def on_cell_changed(self, row, col):
        """Обрабатывает изменение ячейки в таблице."""
        if col == 1:  # Если изменена ячейка "Наименование"
            self.update_code_from_name(row)
        elif col == 6 or col == 7:  # Если изменены "Норма брутто" или "% Потерь"
            self.update_net_weight_and_cost(row)

    def update_code_from_name(self, row):
        """Обновляет код на основе введенного наименования."""
        name_item = self.table.item(row, 1)
        code_item = self.table.item(row, 2)
        if name_item and code_item:
            name = name_item.text()
            if name in self.name_to_code:
                code_item.setText(self.name_to_code[name])

    def update_okei_code(self, row):
        """Обновляет код ОКЕИ на основе выбранной единицы измерения."""
        unit_combo = self.table.cellWidget(row, 3)
        okei_item = self.table.item(row, 4)
        if unit_combo and okei_item:
            unit = unit_combo.currentText()
            if unit in self.unit_to_okei:
                okei_item.setText(self.unit_to_okei[unit])

    def update_net_weight_and_cost(self, row):
        """Обновляет норму нетто и стоимость на основе нормы брутто и процента потерь."""
        try:
            # Получаем значения из ячеек
            gross_weight_item = self.table.item(row, 6)  # Норма брутто
            loss_percent_item = self.table.item(row, 7)  # % Потерь
            price_item = self.table.item(row, 5)  # Цена

            if gross_weight_item and loss_percent_item and price_item:
                gross_weight = float(gross_weight_item.text().replace(',', '.'))
                loss_percent = float(loss_percent_item.text().replace(',', '.'))
                price = float(price_item.text().replace(',', '.'))

                # Расчет нормы нетто
                net_weight = gross_weight * (1 - loss_percent / 100)
                net_weight_item = self.table.item(row, 8)  # Норма нетто
                if not net_weight_item:
                    net_weight_item = QTableWidgetItem()
                    self.table.setItem(row, 8, net_weight_item)
                net_weight_item.setText(f"{net_weight:.2f}")

                # Расчет стоимости
                cost = price * net_weight
                cost_item = self.table.item(row, 9)  # Сумма
                if not cost_item:
                    cost_item = QTableWidgetItem()
                    self.table.setItem(row, 9, cost_item)
                cost_item.setText(f"{cost:.2f}")
        except ValueError:
            pass

    def save_to_xlsx(self, filename):
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active

        ws['N14'] = 'Калькуляционная карточка'

        ws['AS1'] = 'Унифицированная форма № ОП-1'
        ws['AS2'] = 'Утверждена постановлением Госкомстата'
        ws['AS3'] = 'России от 25.12.98 № 132'

        ws['BC4'] = 'Код'
        ws.merge_cells('BC4:BJ4')

        ws['AU5'] = 'Форма по ОКУД'

        ws['AX6'] = 'по ОКПО'

        ws['AP9'] = 'Вид деятельности по ОКДП'

        ws['AJ11'] = 'Номер блюда по сборнику рецептур, ТТК, СТП'

        ws['AV12'] = 'Вид операции'

        ws['BC5'] = '0330501'
        ws.merge_cells('BC5:BJ5')

        ws['BC6'] = ' '
        ws.merge_cells('BC6:BJ6')

        ws['BC7'] = ' '
        ws.merge_cells('BC7:BJ7')

        ws['BC8'] = ' '
        ws.merge_cells('BC8:BJ8')

        ws['BC9'] = ' '
        ws.merge_cells('BC9:BJ9')

        ws['BC10'] = ' '
        ws.merge_cells('BC10:BJ10')

        ws['BC11'] = ' '
        ws.merge_cells('BC11:BJ11')

        operations_codes = self.load_operations_codes()
        selected_operation = self.operation_combobox.currentText()
        operation_code = operations_codes.get(selected_operation,'')

        ws['BC12'] = operation_code
        ws.merge_cells('BC12:BJ12')

        ws['A6'] = self.org_combobox.currentText()
        ws.merge_cells('A6:AU6')

        ws['A8'] = ' '
        ws.merge_cells('A8:AU8')

        ws['A10'] = self.dish_combobox.currentText()
        ws.merge_cells('A10:AU10')

        ws['A7'] = '(организация)'
        ws.merge_cells('A7:AF7')

        ws['A9'] = '(структурное подразделение)'
        ws.merge_cells('A9:AF9')

        ws['A11'] = '(наименование блюда)'
        ws.merge_cells('A11:AF11')

        font_size = 6.5
        font = Font(size=font_size)
        alignment = Alignment(horizontal='center', vertical='center')

        for row in [7, 9, 11]:
            for col in range(1, 32):
                cell = ws.cell(row=row, column=col)
                cell.font = font
                cell.alignment = alignment


        ws['AD13'] = 'Номер документа'
        ws.merge_cells('AD13:AK13')

        ws['AD14'] = self.get_number_value()
        ws.merge_cells('AD14:AK14')

        ws['AL13'] = 'Дата составления'
        ws.merge_cells('AL13:AS13')

        ws['AL14'] = self.get_date()
        ws.merge_cells('AL14:AS14')

        ws['A16'] = ('Порядковый номер калькуляции,'
                     '\n дата утверждения')
        ws.merge_cells('A16:P18')

        alignment = Alignment(horizontal='center', vertical='center')

        ws['A19'] = 'Но-\nмер\nпо по-\nрядку '
        ws.merge_cells('A19:B20')

        ws['C19'] = 'Наименование'
        ws.merge_cells('C19:G20')

        ws['H19'] = 'Код'
        ws.merge_cells('H19:P20')

        ws['Q16'] = 'Единицы измерения'
        ws.merge_cells('Q16:AD18')

        ws['Q19'] = 'Наименование'
        ws.merge_cells('Q19:X20')

        ws['Y19'] = 'Код по ОКЕИ'
        ws.merge_cells('Y19:AD20')

        ws['AE16'] = 'Цена,\n руб. коп.'
        ws.merge_cells('AE16:AL20')

        ws['AM16'] = 'Норма брутто'
        ws.merge_cells('AM16:AT20')

        ws['AU16'] = 'Норма нетто'
        ws.merge_cells('AU16:BB20')

        ws['BC16'] = 'Стоимость,\n руб. коп.'
        ws.merge_cells('BC16:BJ20')

        for row in [16, 17, 18]:  # Ячейки A7, A9 и A11
            for col in range(1, 32):  # Колонки от A (1) до AF (32)
                cell = ws.cell(row=row, column=col)
                cell.alignment = alignment

        font_size = 7
        font = Font(size=font_size)
        alignment = Alignment(horizontal='center', vertical='center')

        for row in [19, 20]:
            for col in range(1, 2):
                cell = ws.cell(row=row, column=col)
                cell.font = font
                cell.alignment = alignment

        # Устанавливаем ширину столбцов для уменьшения расстояния между ними
        column_widths_pt = {
            'A': 8, 'B': 15, 'C': 19, 'D': 30, 'E': 27, 'F': 33, 'G': 33,
            'H': 19, 'I': 12, 'J': 10, 'K': 8, 'L': 14, 'M': 8, 'N': 28, 'O': 8,
            'P': 21, 'Q': 10, 'R': 10, 'S': 8, 'T': 14, 'U': 8, 'V': 28, 'W': 8,
            'X': 21, 'Y': 10, 'Z': 10, 'AA': 8, 'AB': 14, 'AC': 9, 'AD': 27, 'AE': 8, 'AF': 21,
            'AG': 11, 'AH': 10, 'AI': 8, 'AJ': 14, 'AK': 10, 'AL': 15, 'AM': 16, 'AN': 8,
            'AO': 21, 'AP': 11, 'AQ': 10, 'AR': 8, 'AS': 14, 'AT': 8, 'AU': 9, 'AV': 8,
            'AW': 13, 'AX': 9, 'AY': 19, 'AZ': 12, 'BA': 10, 'BB': 12, 'BC': 14, 'BD': 8,
            'BE': 15, 'BF': 15, 'BG': 8, 'BH': 10, 'BI': 10, 'BJ': 12,

        }

        # Устанавливаем ширину столбцов
        for column, pt in column_widths_pt.items():
            ws.column_dimensions[column].width = pt_to_col_width(pt)
        # Определяем, какие столбцы объединены
        column_ranges = [
            ('A', 'B'),  # Первый столбец (№ п/п)
            ('C', 'G'),  # Второй столбец (Наименование)
            ('H', 'P'),  # Третий столбец (Код)
            ('Q', 'X'),  # Четвертый столбец (Единица-Наименование)
            ('Y', 'AD'),  # Пятый столбец (Код ОКЕИ)
            ('AE', 'AL'),  # Шестой столбец (Цена, руб. коп.)
            ('AM', 'AT'),  # Седьмой столбец (Норма брутто)
            ('AU', 'BB'),  # Восьмой столбец (Норма нетто)
            ('BC', 'BJ'),  # Девятый столбец (Сумма, руб. коп.)
        ]

        # Заполняем таблицу данными, начиная с 21-й строки
        start_row = 21  # Начальная строка для данных
        for row in range(self.table.rowCount()):
            row_data = []
            for col in range(self.table.columnCount()):
                if col == 3:  # Столбец "Единица"
                    combo_box = self.table.cellWidget(row, col)
                    row_data.append(combo_box.currentText() if combo_box else "")
                else:
                    item = self.table.item(row, col)
                    row_data.append(item.text() if item else "")

            for col_idx, (start_col, end_col) in enumerate(column_ranges):
                cell = ws[f"{start_col}{start_row + row}"]

                cell.value = row_data[col_idx]

                if start_col != end_col:
                    ws.merge_cells(f"{start_col}{start_row + row}:{end_col}{start_row + row}")

        signatures_row = start_row + self.table.rowCount() + 4
        alignment = Alignment(horizontal="center", vertical="center")
        signature_font = Font(size=6.5)
        ws[f"A{signatures_row}"] = "Заведующий:"

        ws.merge_cells(f"A{signatures_row}:E{signatures_row}")
        ws[f"A{signatures_row+1}"] = ' '
        ws.merge_cells(f"A{signatures_row+1}:L{signatures_row+1}")
        ws[f"M{signatures_row+1}"] = '(подпись)'
        ws[f"M{signatures_row + 1}"].font = signature_font
        ws[f"M{signatures_row + 1}"].alignment = alignment
        ws.merge_cells(f"M{signatures_row + 1}:AC{signatures_row + 1}")

        ws[f"A{signatures_row+2}"] = "Бухгалтер:"

        ws.merge_cells(f"A{signatures_row+2}:E{signatures_row+2}")
        ws[f"A{signatures_row + 3}"] = ' '
        ws.merge_cells(f"A{signatures_row + 3}:L{signatures_row + 3}")
        ws[f"M{signatures_row + 3}"] = '(подпись)'
        ws[f"M{signatures_row + 3}"].font = signature_font
        ws[f"M{signatures_row + 3}"].alignment = alignment
        ws.merge_cells(f"M{signatures_row + 3}:AC{signatures_row + 3}")

        ws[f"A{signatures_row + 4}"] = "Утверждено:"
        ws.merge_cells(f"A{signatures_row+4}:E{signatures_row+4}")

        ws[f"A{signatures_row + 5}"] = ' '
        ws.merge_cells(f"A{signatures_row + 5}:L{signatures_row + 5}")
        ws[f"M{signatures_row + 5}"] = '(подпись)'
        ws[f"M{signatures_row + 5}"].font = signature_font
        ws[f"M{signatures_row + 5}"].alignment = alignment
        ws.merge_cells(f"M{signatures_row + 5}:AC{signatures_row + 5}")


        dialog = SignaturesDialog(self)
        if dialog.exec_() == QDialog.Accepted:
            signatures = dialog.get_signatures()
            ws[f"F{signatures_row}"] = f"{signatures['head']['fio']}"
            ws[f"F{signatures_row + 2}"] = f"{signatures['accountant']['fio']}"
            ws[f"F{signatures_row + 4}"] = f"{signatures['approve']['fio']}"

            ws.merge_cells(f"F{signatures_row}:AC{signatures_row}")
            ws.merge_cells(f"F{signatures_row + 2}:AC{signatures_row + 2}")
            ws.merge_cells(f"F{signatures_row + 4}:AC{signatures_row + 4}")



        # Устанавливаем шрифт и выравнивание для заголовков
        header_font = Font(bold=True)
        for cell in ws[1]:
            cell.font = header_font
            cell.alignment = alignment

        wb.save(filename)


class SaveDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle('Выберите формат сохранения')
        self.setGeometry(150, 150, 300, 150)

        layout = QVBoxLayout()

        self.format_group = QButtonGroup(self)

        pdf_radio = QRadioButton('PDF')
        xlsx_radio = QRadioButton('XLSX')
        self.format_group.addButton(pdf_radio, 1)
        self.format_group.addButton(xlsx_radio, 2)

        layout.addWidget(pdf_radio)
        layout.addWidget(xlsx_radio)

        save_button = QPushButton('Сохранить')
        save_button.clicked.connect(self.save_file)
        layout.addWidget(save_button)

        self.setLayout(layout)

    def save_file(self):
        selected_format = self.format_group.checkedId()
        if selected_format == 1:
            filename, _ = QFileDialog.getSaveFileName(self, "Сохранить как PDF", "", "PDF Files (*.pdf)")
            if filename:
                self.parent().save_to_pdf(filename)
                QMessageBox.information(self, "Успех", "Файл успешно сохранен в формате PDF.")
        elif selected_format == 2:
            filename, _ = QFileDialog.getSaveFileName(self, "Сохранить как XLSX", "", "Excel Files (*.xlsx)")
            if filename:
                self.parent().save_to_xlsx(filename)
                QMessageBox.information(self, "Успех", "Файл успешно сохранен в формате XLSX.")
        self.close()


class PrintDialog(QDialog):
    def __init__(self, data, parent=None):
        super().__init__(parent)
        self.setWindowTitle('Таблица для печати')
        self.setGeometry(200, 200, 1000, 400)

        layout = QVBoxLayout()

        self.table = QTableWidget()
        self.table.setColumnCount(10)
        self.table.setHorizontalHeaderLabels(
            ['№ п/п', 'Наименование', 'Код', 'Единица', 'Код ОКЕИ', 'Цена, руб. коп.', 'Норма брутто', '% Потерь',
             'Норма нетто', 'Сумма, руб. коп.'])

        self.table.setRowCount(len(data))
        for i, row in enumerate(data):
            for j, item in enumerate(row):
                self.table.setItem(i, j, QTableWidgetItem(item))

        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.table.horizontalHeader().setStretchLastSection(True)

        layout.addWidget(self.table)

        close_button = QPushButton('Закрыть')
        close_button.clicked.connect(self.close)
        layout.addWidget(close_button)

        self.setLayout(layout)


if __name__ == '__main__':
    app = QApplication(sys.argv)
    form = OP1Form()
    form.show()
    sys.exit(app.exec_())
